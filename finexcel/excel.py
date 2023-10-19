from email.headerregistry import HeaderRegistry
from openpyxl import *
import string


# List of all letters in the alphabet
single_alphabet_list = list(string.ascii_uppercase)
double_alphabet_list = []  

x=0
for i in range(26):
    y=0
    for i in range(26):
        double_alphabet_list.append(single_alphabet_list[x] + single_alphabet_list[y])
        y+=1
    
    x+=1

final_alphabet_list = single_alphabet_list + double_alphabet_list



def sideBarInformation(AnnualData, ws):
    
    # Titles of the information that will exist in the sidebar
    sideBarNameList = [
        'Name', 
        'Sector', 
        'Industry', 
        'Exchange', 
        '', 
        'Price', 
        '52 Week High', 
        '52 Week Low', 
        'Market Cap', 
        '', 
        'P/E Ratio', 
        'PEG Ratio', 
        'P/S Ratio', 
        'Short Ratio'
        ] #Potential to add: dividends, enterprise value, etc.


    # Data of the information that will exist in the sidebar
    sideBarDataList = [
        AnnualData.getCompanyProfile()[0]['companyName'], 
        AnnualData.getCompanyProfile()[0]['industry'], 
        AnnualData.getCompanyProfile()[0]['sector'], 
        AnnualData.getCompanyProfile()[0]['exchangeShortName'], 
        '', 
        AnnualData.getQuote()[0]['price'], 
        AnnualData.getQuote()[0]['yearHigh'], 
        AnnualData.getQuote()[0]['yearLow'], 
        AnnualData.getQuote()[0]['marketCap'], 
        '', 
        AnnualData.getTTMRatio()[0]['peRatioTTM'], 
        AnnualData.getTTMRatio()[0]['pegRatioTTM'], # Doesn't work well at all
        AnnualData.getTTMRatio()[0]['priceToSalesRatioTTM'], 
        'Currently Unavalible'
        ] 

    # Generating the sidebar information
    x = 0
    for i in range(len(sideBarNameList)):
        ws['A' + str(x+5)] = sideBarNameList[x]
        ws['B' + str(x+5)] = sideBarDataList[x]
        x+=1

    




def mainInformation(AnnualData, ws):
    
    # Generating the years of the titles

    minus_years = 0 # amount of years subtracting
    num_year = 20 # number of loops/years back I want to go 
    alpha_pos = 25 # The list positon of last letter in the alphabet
    
    year_date = str(AnnualData.getIncomeStatement()[0]['date']).split('-') #getting list of date broken down in [year, month, date]

    for i in range(num_year):
        try:
            # Using split() to get only the year
            ws[final_alphabet_list[alpha_pos] + '4'] = int(year_date[0]) - minus_years
        
        except(Exception):
            ws[final_alphabet_list[alpha_pos] + '4'] = 'Error'
        
        minus_years += 1
        alpha_pos -= 1

    # Generating the titles
    
    indictors = [
        'Period End',
        'Shares Outstanding',
        'Revenues',
        'Revenue 3 year Average',
        'Turnover',
        'Turnover 3 year Average', # Divided by something
        'Net Income', # Divided by the same thing as above
        'ROIC', #Return on invested Capital
        'ROIC 3 year average',
        'Gross Profit Margin',
        'Gross Profit Margin 3 year Average',
        'EBITDA Margin',
        'Net Income Margin',
        'CFO', #Cash flow from operations
        'CFI', #Cash from investing activities
        'CFF', #Cash from financing activities
        'NCF', #Net Cash flow
        'FCF', #free cash flow
        'Revenue per share',
        'Assets per share',
        'Book value per share',
        'Tangible book value per share',
        'Return on Equity',
        'Earning per share',
        'Earning per share 3 year average',
        'EBITDA per share',
        'EBITDA per share 3 year average',
        'Net income per share',
        'Net income per share 3 year average',
        'CFO per share',
        'CFO per share 3 year average',
        'FCF per share',
        'FCF per share 3 year average'
    ]

    x = 0
    for i in range(len(indictors)):
        ws['F' + str(x+5)] = indictors[x]
        x+=1

    
    # Generating the information

    #All the data needs to be written as a string as we will be looping it multiple times (through an eval function)
    data = [
        AnnualData.getIncomeStatement()[x]['date'], # Date
        AnnualData.getEnterpriseValue()[x]['numberOfShares'], # Shares Outstanding
        AnnualData.getIncomeStatement()[x]['revenue'], # Revenue
        (AnnualData.getIncomeStatement()[x]['revenue'] + AnnualData.getIncomeStatement()[x+1]['revenue'] + AnnualData.getIncomeStatement()[x+2]['revenue']) / 3, # Revenue 3 year average
        AnnualData.getCompanyRatio()[x]['assetTurnover'], # Asset Turnover
        (AnnualData.getCompanyRatio()[x]['assetTurnover'] + AnnualData.getCompanyRatio()[x+1]['assetTurnover'] + AnnualData.getCompanyRatio()[x+2]['assetTurnover']) / 3, # Asset Turnover 3 year average
        AnnualData.getIncomeStatement()[x]['netIncome'], #net income
        AnnualData.getKeyMetrics()[x]['roic'], # Roic
        (AnnualData.getKeyMetrics()[x]['roic'] + AnnualData.getKeyMetrics()[x+1]['roic'] + AnnualData.getKeyMetrics()[x+2]['roic']) / 3, # Roic 3 year average
        AnnualData.getCompanyRatio()[x]['grossProfitMargin'], # Gross profit margin
        (AnnualData.getCompanyRatio()[x]['grossProfitMargin'] + AnnualData.getCompanyRatio()[x+1]['grossProfitMargin'] + AnnualData.getCompanyRatio()[x+2]['grossProfitMargin']) / 3, # Gross profit margin 3 year average
        AnnualData.getIncomeStatement()[x]['ebitdaratio'], # EBITDA Margin or EBITDA Ratio
        AnnualData.getIncomeStatement()[x]['ebitdaratio'], #Net Income Margin
        AnnualData.getCashFlow()[x]['operatingCashFlow'], # Cash flow from operations
        AnnualData.getCashFlow()[x]['netCashUsedForInvestingActivites'], # Cash flow from investing activities
        AnnualData.getCashFlow()[x]['netCashUsedProvidedByFinancingActivities'], # Cash flow from financing activies
        AnnualData.getCashFlow()[x]['netChangeInCash'], # net cash flow
        AnnualData.getCashFlow()[x]['freeCashFlow'], # free cash flow
        AnnualData.getKeyMetrics()[x]['revenuePerShare'], # Revenue per share
        AnnualData.getIncomeStatement()[x]['ebitdaratio'], # Asset per share
        AnnualData.getKeyMetrics()[x]['bookValuePerShare'], # book value per share
        AnnualData.getKeyMetrics()[x]['tangibleBookValuePerShare'], # tangible book value per share
        AnnualData.getKeyMetrics()[x]['roe'], # return on equity
        AnnualData.getIncomeStatement()[x]['eps'], # earning per share
        (AnnualData.getIncomeStatement()[x]['eps'] + AnnualData.getIncomeStatement()[x+1]['eps'] + AnnualData.getIncomeStatement()[x+2]['eps']) / 3, # earning per share 3 year average
        (AnnualData.getIncomeStatement()[x]['ebitda']) / (AnnualData.getEnterpriseValue()[x]['numberOfShares']), # Ebitda per share
        (((AnnualData.getIncomeStatement()[x]['ebitda']) / (AnnualData.getEnterpriseValue()[x]['numberOfShares'])) + ((AnnualData.getIncomeStatement()[x+1]['ebitda']) / (AnnualData.getEnterpriseValue()[x+1]['numberOfShares'])) + ((AnnualData.getIncomeStatement()[x+2]['ebitda']) / (AnnualData.getEnterpriseValue()[x+2]['numberOfShares']))) / 3, # ebitda per share 3 year average
        AnnualData.getKeyMetrics()[x]['netIncomePerShare'],
        (AnnualData.getKeyMetrics()[x]['netIncomePerShare'] + AnnualData.getKeyMetrics()[x+1]['netIncomePerShare'] + AnnualData.getKeyMetrics()[x+2]['netIncomePerShare']) / 3,
        AnnualData.getKeyMetrics()[x]['operatingCashFlowPerShare'], # operating cash flow per share
        (float(AnnualData.getKeyMetrics()[x]['operatingCashFlowPerShare'] or 0) + float(AnnualData.getKeyMetrics()[x+1]['operatingCashFlowPerShare'] or 0) + float(AnnualData.getKeyMetrics()[x+2]['operatingCashFlowPerShare'] or 0)) / 3, # operating cash flow per share 3 year average
        AnnualData.getKeyMetrics()[x]['freeCashFlowPerShare'], # free cash flow per share
        (float(AnnualData.getKeyMetrics()[x]['freeCashFlowPerShare'] or 0) + float(AnnualData.getKeyMetrics()[x+1]['freeCashFlowPerShare'] or 0) + float(AnnualData.getKeyMetrics()[x+2]['freeCashFlowPerShare'] or 0)) / 3 # free cash flow per share 3 year average
    ]
   
            
    # This loops moves across than 1 done until its reachs the end
    expression = 0
    starting_row = 5
    for name in indictors: # "2" should be len(indicators)   
        x = 0
        alpha_pos = 25 # This value should also be the same 
        
        try:
            for i in range(num_year):
                ws[final_alphabet_list[alpha_pos] + str(starting_row)] = data[expression] #starts at the end of the row then comes in reverse
                x += 1
                alpha_pos -= 1
        
        
        except(Exception):
            pass
            

        
        expression += 1 
        starting_row += 1


    
            
        


