from excel import mainInformation, sideBarInformation 
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
import string
from openpyxl import *
from info import FundmentalData

alphabet_list = list(string.ascii_uppercase)


def formatNumbers(ws):
    # Large numbers (accounting style)
    number_format_rows = [6]
    x=0
    for i in range(len(number_format_rows)):
        y=6
        for i in range(20):
            ws[alphabet_list[y] + str(number_format_rows[x])].number_format = '''_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'''
            y+=1
        x+=1

    # Large numbers with dollar signs (accounting style)
    number_format_rows = [7,8,11,18,19,20,21,22]
    x=0
    for i in range(len(number_format_rows)):
        y=6
        for i in range(20):
            ws[alphabet_list[y] + str(number_format_rows[x])].number_format = '''_-$* #,##0_-;-$* #,##0_-;_-$* "-"_-;_-@_-'''
            y+=1
        x+=1

    # Percentages (accounting style)
    number_format_rows = [12,13,14,15,16]
    x=0
    for i in range(len(number_format_rows)):
        y=6
        for i in range(20):
            ws[alphabet_list[y] + str(number_format_rows[x])].number_format = '''_-* #,##0.00%_-;-* #,##0.00%_-;_-* "-"??_-;_-@_-'''
            y+=1
        x+=1

    # Smaller numbers with decimals (accounting style)
    number_format_rows = [9,10,23,25,26,27,28,29,30,31,32,33,34,35,36,37]
    x=0
    for i in range(len(number_format_rows)):
        y=6
        for i in range(20):
            ws[alphabet_list[y] + str(number_format_rows[x])].number_format = '''_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-'''
            y+=1
        x+=1


def autoSizeColumns(ws):
    for col in ws.columns:
         max_length = 0
         column = col[0].column_letter # Get the column name

         for cell in col:
             try: # Necessary to avoid error on empty cells
                 if len(str(cell.value)) > max_length:
                     max_length = len(str(cell.value))

             except(Exception):
                 print(Exception)

         adjusted_width = (max_length + 2)
         ws.column_dimensions[column].width = adjusted_width



def formatGeneralInfo(ws):
    # Aligning general information
    x = 1
    for i in range(50): # Sloppy to just put 50
        ws[alphabet_list[0] + str(x)].alignment = Alignment(horizontal="left")
        ws[alphabet_list[1] + str(x)].alignment = Alignment(horizontal="left")
        x+=1

    # Borders on general information
    number_format_rows = [5,6,7,8,10,11,12,13,15,16,17,18] #These are the rows in the general information section that will have a border
    
    x=0
    for i in range(len(number_format_rows)):
        ws["A"+ str(number_format_rows[x])].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws["B"+ str(number_format_rows[x])].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        x+=1

    #Bolding titles on general information
    x=0
    for i in range(len(number_format_rows)):
        ws["A"+ str(number_format_rows[x])].font = Font(bold=True)
        x+=1

def formatData(ws): 
    x_length = 21
    y_length = 34
    start_offset = 4
    
    # Adding borders around data 
    for i in range(x_length):
        for j in range(y_length): 
            ws[alphabet_list[i + 5] + str(j + start_offset)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Bolding the first column and top row
    for i in range(x_length):
        ws[alphabet_list[i + 5] + str(start_offset)].font = Font(bold=True)
     
    for i in range(y_length):
        ws[alphabet_list[5] + str(i + start_offset)].font = Font(bold=True)




def generateSpreadsheet(ticker, api_key):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "fundementalData"

    AnnualData = FundmentalData(ticker, api_key, 'annual') 
    # QuarterlyData = FundmentalData(ticker, api_key, 'quarter')

    # Getting Financial Info 
    sideBarInformation(AnnualData, ws)
    mainInformation(AnnualData, ws)

    # Formatting Financial Info
    formatNumbers(ws)
    autoSizeColumns(ws)
    formatGeneralInfo(ws)
    formatData(ws)
    
    wb.save('../spreadsheets/' + ticker + '.xlsx')





    